import streamlit as st
import pdfplumber, re, pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, tempfile

st.set_page_config(page_title="Extracteur PDF vers Excel", page_icon="📄", layout="centered")

st.markdown("""
<style>
.stApp { background: linear-gradient(160deg, #6B0000 0%, #8B0000 40%, #B22222 100%); }
.main-title { color:#FFFFFF; font-size:2.2em; font-weight:800; text-align:center; text-shadow:1px 2px 6px rgba(0,0,0,0.4); margin-bottom:5px; }
.sub-title { color:#FFCDD2; font-size:1em; text-align:center; margin-bottom:20px; }
.badge-container { text-align:center; margin-bottom:25px; }
.badge { display:inline-block; background:rgba(255,255,255,0.15); color:#FFFFFF; border:1px solid rgba(255,255,255,0.4); padding:4px 14px; border-radius:20px; font-size:0.82em; margin:3px; }
.card { background:#FFFFFF; border-radius:16px; padding:24px; box-shadow:0 8px 30px rgba(0,0,0,0.25); margin:10px 0; }
.section-label { color:#8B0000; font-weight:700; font-size:0.82em; text-transform:uppercase; letter-spacing:1px; border-left:3px solid #8B0000; padding-left:8px; margin-bottom:10px; }
.stButton > button { background:linear-gradient(135deg,#8B0000,#C0392B) !important; color:white !important; border:none !important; border-radius:10px !important; font-size:1em !important; font-weight:700 !important; width:100% !important; }
.success-box { background:#E8F5E9; border:1px solid #4CAF50; border-radius:10px; padding:15px; color:#1B5E20; font-weight:600; }
.error-box { background:#FFEBEE; border:1px solid #F44336; border-radius:10px; padding:15px; color:#B71C1C; font-weight:600; }
</style>
""", unsafe_allow_html=True)

COLONNES_SORTIE = ["Txn. Date", "Value Date", "Description", "Ex-ref no", "Txn. Ref No", "Debit", "Credit", "Balance"]
MOTIF_DATE_F1 = re.compile(r"^\d{2}-\d{2}-\d{4}$")
MOTIF_DATE_F2 = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")
MOTIF_DATE_F3 = re.compile(r"^\d{4}-\d{2}-\d{2}$")
FOOTER_KEYWORDS = {"computer-generated", "signature.", "Page"}
MAPPING = {
    "date":"Txn. Date","txn. date":"Txn. Date","txn.date":"Txn. Date",
    "value date":"Value Date","description":"Description","narration":"Description",
    "particulars":"Description","check no":"Txn. Ref No","cheque no":"Txn. Ref No",
    "txn. ref no":"Txn. Ref No","ref no":"Txn. Ref No","ex-ref no":"Ex-ref no",
    "debit":"Debit","debits":"Debit","debit (usk)":"Debit","debits (usk)":"Debit",
    "credit":"Credit","credits":"Credit","credit (usk)":"Credit","credits (usk)":"Credit",
    "amount":"Credit","balance":"Balance","branch":None,
}

def _est_date(texte):
    for fmt, motif in [("%d-%m-%Y",MOTIF_DATE_F1),("%d-%b-%Y",MOTIF_DATE_F2),("%Y-%m-%d",MOTIF_DATE_F3)]:
        if motif.match(texte):
            try: datetime.strptime(texte, fmt); return True
            except ValueError: pass
    try: datetime.strptime(texte[:10], "%Y-%m-%d"); return True
    except ValueError: pass
    return False

def _est_montant(t): return bool(re.match(r"^\d[\d,]*\.?\d{0,2}$", t.replace(" ","")))

def _parse(v):
    try: return float(str(v).replace(",","").replace(" ",""))
    except: return None

def _extraire_entete(chemin_pdf):
    info = {k:"" for k in ["banque","customer_name","customer_address","period_from","period_to","account_number","customer_number","currency","account_open_date","bic_code"]}
    with pdfplumber.open(chemin_pdf) as pdf:
        mots = pdf.pages[0].extract_words(x_tolerance=3, y_tolerance=3)
        if not mots: return info
        t = " ".join(m["text"] for m in sorted(mots, key=lambda m:(m["top"],m["x0"])))
        for nom in ["EXIM BANK","Exim Bank","BANQUE NATIONALE DE DJIBOUTI","BANQUE NATIONALE"]:
            if nom.upper() in t.upper(): info["banque"]=nom; break
        for pattern in [
            r"Account\s+Statement\s+from\s+(\d{2}-[A-Za-z]{3}-\d{4})\s+to\s+(\d{2}-[A-Za-z]{3}-\d{4})",
            r"Period[:\s]+(\d{2}/\d{2}/\d{4})\s+to\s+(\d{2}/\d{2}/\d{4})",
        ]:
            m = re.search(pattern, t, re.IGNORECASE)
            if m: info["period_from"],info["period_to"]=m.group(1),m.group(2); break
        m = re.search(r"Customer\s+Name\s+&\s+Address\s+(.*?)\s+Branch\s+Name\s+&\s+Address", t)
        if m:
            wc=m.group(1).strip().split(); nom,addr,ia=[],[],False
            for w in wc:
                if w.upper() in ("DJIBOUTI","SOMALIA","ETHIOPIA") or ia: ia=True; addr.append(w)
                else: nom.append(w)
            info["customer_name"],info["customer_address"]=" ".join(nom)," ".join(addr)
        for pattern,key in [
            (r"Account\s+Number[:\s]+(\S+)","account_number"),
            (r"Account\s+No[:\s]+(\S+)","account_number"),
            (r"Customer\s+Number[:\s]+(\S+)","customer_number"),
            (r"Account\s+Open\s+Date[:\s]+([A-Za-z]+\s+\d{1,2},?\s+\d{4})","account_open_date"),
            (r"BIC\s+Code[:\s]+(\S+)","bic_code"),
        ]:
            mm=re.search(pattern,t,re.IGNORECASE)
            if mm and not info[key]: info[key]=mm.group(1).strip()
        for pattern in [
            r"Currency[:\s]+(DJF|USD|EUR|GBP|ETB)\s*[-]\s*([\w\s]+?)(?=Account|BIC|$)",
            r"Currency[:\s]+(DJF|USD|EUR|GBP|ETB)",r"\b(USD|DJF|EUR|GBP|ETB)\b",
        ]:
            mm=re.search(pattern,t,re.IGNORECASE)
            if mm:
                info["currency"]=mm.group(1)
                if len(mm.groups())>1 and mm.group(2): info["currency"]+=" - "+mm.group(2).strip()
                break
    return info

def _extraire_transactions(chemin_pdf):
    resultats = []
    # Methode 1 : par tableau
    with pdfplumber.open(chemin_pdf) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables({"vertical_strategy":"lines","horizontal_strategy":"lines"}):
                if not table or len(table)<2: continue
                hr,hi=None,0
                for idx,row in enumerate(table):
                    if row and any(c and re.search(r"date|description|debit|credit|balance",str(c),re.IGNORECASE) for c in row):
                        hr=[str(c).strip().lower() if c else "" for c in row]; hi=idx; break
                if not hr: continue
                for row in table[hi+1:]:
                    if not row or not row[0]: continue
                    if not _est_date(str(row[0]).strip()[:19]): continue
                    r={c:"" for c in COLONNES_SORTIE}
                    for ci,cn in enumerate(hr):
                        if ci<len(row):
                            cc=MAPPING.get(cn)
                            if cc and row[ci]: r[cc]=str(row[ci]).replace("\n"," ").strip()
                    resultats.append(r)
    if resultats: return resultats
    # Methode 2 : par mots
    trans_brutes = []
    with pdfplumber.open(chemin_pdf) as pdf:
        for page in pdf.pages:
            mots=page.extract_words(x_tolerance=3,y_tolerance=3)
            if not mots: continue
            mt=sorted(mots,key=lambda m:(m["top"],m["x0"]))
            yf=float("inf")
            for mot in mt:
                if mot["text"] in FOOTER_KEYWORDS: yf=mot["top"]; break
            pa=[]
            for mot in mt:
                if _est_date(mot["text"][:19]) and mot["x0"]<150: pa.append({"date":mot["text"],"y":mot["top"]})
            for i,a in enumerate(pa):
                yd=a["y"]; ye=pa[i+1]["y"] if i+1<len(pa) else yf
                tran=[m for m in mt if yd<=m["top"]<ye]
                if tran: trans_brutes.append((a["date"],tran))
    for idx,(date,tran) in enumerate(trans_brutes):
        mb=[m["text"] for m in tran]; txr=""
        for mo in tran:
            if re.match(r"^\d+$",mo["text"]) and 255<=mo["x0"]<=330: txr=mo["text"]; break
        if not txr: txr=next((m for m in mb if re.match(r"^\d+/\d+/\d+$",m)),"")
        mons,desc=[],[]
        for mo in tran[1:]:
            tt=mo["text"]
            if tt==txr: continue
            if _est_date(tt[:19]): continue
            if _est_montant(tt) and mo["x0"]>300: mons.append(tt)
            elif mo["x0"]<300 and tt not in ("Main","Branch","Main Branch"): desc.append(tt)
        bal=mons[-1] if mons else ""; amt=mons[-2] if len(mons)>=2 else (mons[0] if mons else "")
        mn=_parse(amt); bc=_parse(bal)
        bp=_parse(resultats[-1].get("Balance","")) if resultats else None
        db=cr=""
        if mn is not None:
            if bc is not None and bp is not None:
                if bc>=bp: cr=amt
                else: db=amt
            else: cr=amt
        resultats.append({"Txn. Date":date,"Value Date":"","Description":" ".join(desc),"Ex-ref no":"","Txn. Ref No":txr,"Debit":db,"Credit":cr,"Balance":bal})
    return resultats

def _sauvegarder_excel(info, transactions, chemin_excel):
    wb=Workbook(); ws=wb.active; ws.title="Releve Bancaire"
    fb=PatternFill("solid",fgColor="1F4E79"); fg=PatternFill("solid",fgColor="D9E1F2")
    fc=PatternFill("solid",fgColor="EBF3FB"); fh=PatternFill("solid",fgColor="2E75B6"); fa=PatternFill("solid",fgColor="FFFFFF")
    bord=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    ac=Alignment(horizontal="center",vertical="center",wrap_text=True); al=Alignment(horizontal="left",vertical="center",wrap_text=True)
    nc=len(COLONNES_SORTIE)
    def es(l,t):
        ws.merge_cells(f"A{l}:{get_column_letter(nc)}{l}"); c=ws.cell(row=l,column=1,value=t)
        c.font=Font(bold=True,size=11,color="FFFFFF"); c.fill=fb; c.alignment=ac; ws.row_dimensions[l].height=20; return l+1
    def ec(l,label,val):
        if not val: return l
        cl=ws.cell(row=l,column=1,value=label); cv=ws.cell(row=l,column=2,value=val)
        ws.merge_cells(f"B{l}:{get_column_letter(nc)}{l}")
        cl.font=Font(bold=True); cl.fill=fg; cl.alignment=al; cl.border=bord
        cv.fill=fc; cv.alignment=al; cv.border=bord; ws.row_dimensions[l].height=16; return l+1
    ligne=1
    if info.get("banque"): ligne=es(ligne,info["banque"])
    if info.get("customer_name"):
        ligne=es(ligne,"Customer Name & Address")
        ligne=ec(ligne,"Customer Name",info.get("customer_name",""))
        ligne=ec(ligne,"Address",info.get("customer_address","")); ligne+=1
    champs=[
        ("Periode",(info.get("period_from","")+" to "+info.get("period_to","")) if info.get("period_from") else ""),
        ("Account No",info.get("account_number","")),("Customer Number",info.get("customer_number","")),
        ("Currency",info.get("currency","")),("Account Open Date",info.get("account_open_date","")),
        ("BIC Code",info.get("bic_code","")),
    ]
    if any(v for _,v in champs):
        for label,val in champs: ligne=ec(ligne,label,val)
        ligne+=1
    for ci,cn in enumerate(COLONNES_SORTIE,start=1):
        c=ws.cell(row=ligne,column=ci,value=cn); c.font=Font(bold=True,color="FFFFFF"); c.fill=fh; c.alignment=ac; c.border=bord
    ws.row_dimensions[ligne].height=18; ligne+=1
    for i,tran in enumerate(transactions):
        rf=fc if i%2==0 else fa
        for ci,cn in enumerate(COLONNES_SORTIE,start=1):
            c=ws.cell(row=ligne,column=ci,value=tran.get(cn,"")); c.fill=rf; c.alignment=al; c.border=bord
        ws.row_dimensions[ligne].height=15; ligne+=1
    larg={"Txn. Date":16,"Value Date":16,"Description":50,"Ex-ref no":14,"Txn. Ref No":16,"Debit":14,"Credit":14,"Balance":16}
    for ci,cn in enumerate(COLONNES_SORTIE,start=1):
        ws.column_dimensions[get_column_letter(ci)].width=larg.get(cn,16)
    wb.save(chemin_excel)

def extraire_pdf_vers_excel(chemin_pdf, chemin_excel):
    try:
        info=_extraire_entete(chemin_pdf)
        transactions=_extraire_transactions(chemin_pdf)
        if not transactions: return False,"Aucune transaction trouvee. Verifiez que le PDF est electronique."
        _sauvegarder_excel(info,transactions,chemin_excel)
        return True,f"{len(transactions)} transactions extraites ({info.get('banque','OK')})"
    except Exception as e: return False,f"Erreur : {str(e)}"

# ── Interface ─────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">Extracteur PDF vers Excel</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Convertissez vos releves bancaires en fichier Excel en un seul clic</div>', unsafe_allow_html=True)
st.markdown('<div class="badge-container"><span class="badge">Banque Nationale de Djibouti</span><span class="badge">Exim Bank</span><span class="badge">Tout PDF electronique</span></div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown('<div class="card"><div class="section-label">Etape 1 - Votre releve PDF</div>', unsafe_allow_html=True)
    fichier_pdf = st.file_uploader("", type=["pdf"], label_visibility="collapsed")
    convertir = st.button("Lancer la conversion")
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="card"><div class="section-label">Etape 2 - Resultat et Telechargement</div>', unsafe_allow_html=True)
    if convertir:
        if fichier_pdf is None:
            st.markdown('<div class="error-box">Veuillez uploader un fichier PDF.</div>', unsafe_allow_html=True)
        else:
            with st.spinner("Conversion en cours..."):
                tmp_dir=tempfile.mkdtemp()
                chemin_pdf=os.path.join(tmp_dir,fichier_pdf.name)
                with open(chemin_pdf,"wb") as f: f.write(fichier_pdf.read())
                nom_excel=fichier_pdf.name.replace(".pdf","_converti.xlsx")
                chemin_excel=os.path.join(tmp_dir,nom_excel)
                succes,message=extraire_pdf_vers_excel(chemin_pdf,chemin_excel)
            if succes:
                st.markdown(f'<div class="success-box">{message}</div>', unsafe_allow_html=True)
                with open(chemin_excel,"rb") as f:
                    st.download_button(
                        label="Telecharger l Excel",
                        data=f,
                        file_name=nom_excel,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.markdown(f'<div class="error-box">{message}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<p style="text-align:center;color:rgba(255,255,255,0.5);font-size:0.8em;margin-top:20px">Le PDF doit etre un document electronique - pas une photo ou un scan.</p>', unsafe_allow_html=True)
